import openpyxl as xl
from openpyxl.chart import BarChart, Reference



# 讀取excel
wb = xl.load_workbook('/Users/Jean/PycharmProjects/HelloWorld/materials/transactions.xlsx')
# 讀取excel的第一張sheet
sheet = wb['Sheet1']
# sheet a1表格
cell = sheet['a1']
# cell = sheet.cell(1, 1)
print(cell.value)

print('-----------------')

# 有幾row
print(sheet.max_row)

print('-----------------')

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # 代表欄位3每列的值
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)  # 代表欄位4每列的值
    corrected_price_cell.value = corrected_price

values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('/Users/Jean/PycharmProjects/HelloWorld/materials/transactions2.xlsx')

